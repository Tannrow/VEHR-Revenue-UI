from __future__ import annotations

from dataclasses import dataclass
import math
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional, Sequence

import pandas as pd
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class TableFrame:
    name: str
    df: pd.DataFrame


def _safe_len(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, float) and math.isnan(value):
        return 0
    try:
        if pd.isna(value):
            return 0
    except Exception:
        pass
    return len(str(value))


def write_claim_lines_xlsx(out_path: Path, rows: Sequence[dict[str, Any]]) -> None:
    """Write a single-sheet workbook named 'ClaimLines' with the expected columns."""
    cols = [
        "Patient Name",
        "Member ID",
        "Claim Line ID",
        "Dates of Service",
        "Modifier/Units",
        "Charge",
        "Payment",
    ]
    df = pd.DataFrame(list(rows))
    # Enforce column order and ensure columns exist even if empty.
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    df = df[cols]

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ClaimLines")
        ws = writer.book["ClaimLines"]
        ws.freeze_panes = "A2"
        for col_idx, col in enumerate(cols, start=1):
            series = df[col].astype(str)
            max_len = max([_safe_len(col)] + [_safe_len(v) for v in series.head(500).tolist()])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)


def write_recon_lines_xlsx(
    out_path: Path,
    rows: Sequence[dict[str, Any]],
    *,
    sheet_name: str = "ReconLines",
    meta: dict[str, Any] | None = None,
) -> None:
    """Write a single-sheet workbook for content-based ERA parsing."""
    cols = [
        "account_id",
        "payer_claim_number",
        "icn",
        "patient_name",
        "member_id",
        "claim_id",
        "line_ctrl_number",
        "dos_from",
        "dos_to",
        "proc_code",
        "units",
        "billed_amount",
        "allowed_amount",
        "paid_amount",
        "adj_code",
        "adj_amount",
        "era_row_confidence",
        "row_is_probably_junk",
        "source_layout",
    ]
    df = pd.DataFrame(list(rows))
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    df = df[cols]

    out_path.parent.mkdir(parents=True, exist_ok=True)
    distinct_patient_name_in_rows = len({row.get("patient_name") for row in rows if row.get("patient_name")})
    distinct_member_id_in_rows = len({row.get("member_id") for row in rows if row.get("member_id")})
    distinct_patient_name_in_df = int(df["patient_name"].nunique(dropna=True)) if "patient_name" in df.columns else 0
    distinct_member_id_in_df = int(df["member_id"].nunique(dropna=True)) if "member_id" in df.columns else 0

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"
        for col_idx, col in enumerate(cols, start=1):
            series = df[col].astype(str)
            max_len = max([_safe_len(col)] + [_safe_len(v) for v in series.head(500).tolist()])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

        if meta is not None and sheet_name == "era_lines":
            meta_rows = []
            generated_at = datetime.now(timezone.utc).isoformat()
            meta_payload = {"generated_at_utc": generated_at}
            meta_payload["distinct_patient_name_in_rows_before_df"] = distinct_patient_name_in_rows
            meta_payload["distinct_member_id_in_rows_before_df"] = distinct_member_id_in_rows
            meta_payload["distinct_patient_name_in_written_sheet"] = distinct_patient_name_in_df
            meta_payload["distinct_member_id_in_written_sheet"] = distinct_member_id_in_df
            meta_payload.update(meta)
            for key, value in meta_payload.items():
                meta_rows.append({"key": key, "value": value})
            meta_df = pd.DataFrame(meta_rows)
            meta_df.to_excel(writer, index=False, sheet_name="_meta")


def _safe_sheet_name(name: str) -> str:
    # Excel: max 31 chars, cannot contain: : \ / ? * [ ]
    bad = [":", "\\", "/", "?", "*", "[", "]"]
    for ch in bad:
        name = name.replace(ch, " ")
    name = " ".join(name.split()).strip()
    return (name or "Table")[:31]


def write_tables_to_xlsx(out_path: Path, tables: list[TableFrame]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        if not tables:
            pd.DataFrame([{"message": "No tables detected in this document."}]).to_excel(
                writer, index=False, sheet_name="NoTables"
            )
            return

        used_names: set[str] = set()
        for i, t in enumerate(tables, start=1):
            base = _safe_sheet_name(t.name or f"Table {i}")
            sheet = base
            n = 2
            while sheet in used_names:
                suffix = f" {n}"
                sheet = _safe_sheet_name(base[: max(0, 31 - len(suffix))] + suffix)
                n += 1
            used_names.add(sheet)

            t.df.to_excel(writer, index=False, sheet_name=sheet)

            ws = writer.book[sheet]
            ws.freeze_panes = "A2"

            # Auto-size columns (approx) with a cap to keep it readable.
            for col_idx, col in enumerate(t.df.columns, start=1):
                series = t.df[col].astype(str) if col in t.df.columns else pd.Series([""])
                max_len = max([_safe_len(col)] + [_safe_len(v) for v in series.head(200).tolist()])
                width = min(max(10, max_len + 2), 60)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
