from __future__ import annotations

import argparse
import logging
import os
import re
import time
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import Session

from app.core.time import utc_now
from app.db.models.billed_line import BilledLine
from app.db.models.era_line import EraLine
from app.db.models.recon_claim_result import ReconClaimResult
from app.db.models.recon_import_job import ReconImportJob
from app.db.models.recon_line_result import ReconLineResult
from app.db.models.user import User
from app.db.session import SessionLocal
from app.services.audit import log_event
from scripts.era_extract.content_parsers import parse_billed_content, parse_era_content
from scripts.era_extract.docintel_client import create_document_intelligence_client, load_repo_dotenv
from app.db.models.document_analysis import DocumentAnalysis, DocumentType


logger = logging.getLogger(__name__)

POLL_SECONDS = 10
CLAIM_LIMIT = 5
RECON_TOLERANCE = Decimal("0.01")


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value)).date()
    except Exception:
        return None


def _sanitize_error_message(exc: Exception) -> str:
    raw = f"{exc.__class__.__name__}: {str(exc)}"
    cleaned = re.sub(r"[A-Za-z0-9]{16,}", "[redacted]", raw)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned[:240]


def _safe_add(a: Decimal | None, b: Decimal | None) -> Decimal:
    return (a or Decimal("0")) + (b or Decimal("0"))


def _norm_text(value: str | None) -> str:
    return (value or "").strip().upper()


def _claim_id(row: dict[str, Any]) -> str | None:
    return (row.get("claim_id") or row.get("account_id") or "").strip() or None


def _same_service(billed_row: dict[str, Any], era_row: dict[str, Any], tolerance: Decimal) -> bool:
    billed_proc = _norm_text(billed_row.get("proc_code"))
    era_proc = _norm_text(era_row.get("proc_code"))
    if billed_proc and era_proc and billed_proc != era_proc:
        return False
    billed_dos = billed_row.get("dos_from")
    era_dos = era_row.get("dos_from")
    if billed_dos is not None and era_dos is not None and billed_dos != era_dos:
        return False

    billed_amt = _to_decimal(billed_row.get("billed_amount"))
    era_amt = _to_decimal(era_row.get("billed_amount"))
    if billed_amt is not None and era_amt is not None and abs(billed_amt - era_amt) > tolerance:
        return False
    return True


def _is_denial_code(code: str | None) -> bool:
    normalized = _norm_text(code)
    if not normalized:
        return False
    if normalized in {"CO-45"}:
        return False
    if normalized.startswith("PR-"):
        return True
    if normalized in {"CO-16", "CO-18", "CO-50", "CO-96", "CO-197", "PI-204"}:
        return True
    return False


@dataclass
class ReconSummary:
    matched_claims: int = 0
    unmatched_era_claims: int = 0
    unmatched_billed_claims: int = 0
    underpaid_claims: int = 0
    denied_claims: int = 0
    needs_review_claims: int = 0


def _reconcile_rows(
    *,
    era_rows: list[dict[str, Any]],
    billed_rows: list[dict[str, Any]],
    tolerance: Decimal,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], ReconSummary]:
    claim_results: list[dict[str, Any]] = []
    line_results: list[dict[str, Any]] = []
    summary = ReconSummary()

    billed_by_claim: dict[str, list[dict[str, Any]]] = defaultdict(list)
    era_by_claim: dict[str, list[dict[str, Any]]] = defaultdict(list)
    billed_missing: list[dict[str, Any]] = []
    era_missing: list[dict[str, Any]] = []

    for row in billed_rows:
        claim_id = _claim_id(row)
        if claim_id:
            billed_by_claim[claim_id].append(row)
        else:
            billed_missing.append(row)
    for row in era_rows:
        claim_id = _claim_id(row)
        if claim_id:
            era_by_claim[claim_id].append(row)
        else:
            era_missing.append(row)

    all_claims = sorted(set(billed_by_claim.keys()) | set(era_by_claim.keys()))
    for claim_id in all_claims:
        billed_group = billed_by_claim.get(claim_id, [])
        era_group = era_by_claim.get(claim_id, [])

        used_era: set[int] = set()
        collisions = 0
        for billed in billed_group:
            candidates = [
                idx
                for idx, era in enumerate(era_group)
                if idx not in used_era and _same_service(billed, era, tolerance)
            ]
            if len(candidates) == 1:
                idx = candidates[0]
                used_era.add(idx)
                era = era_group[idx]
                billed_amount = _to_decimal(billed.get("billed_amount"))
                paid_amount = _to_decimal(era.get("paid_amount"))
                variance = None
                if billed_amount is not None and paid_amount is not None:
                    variance = paid_amount - billed_amount
                status = "matched"
                reason_code = era.get("adj_code")
                if variance is not None and abs(variance) > tolerance:
                    status = "needs_review"
                    reason_code = "variance_outside_tolerance"
                line_results.append(
                    {
                        "account_id": claim_id,
                        "claim_id": claim_id,
                        "dos_from": billed.get("dos_from") or era.get("dos_from"),
                        "dos_to": billed.get("dos_to") or era.get("dos_to"),
                        "proc_code": billed.get("proc_code") or era.get("proc_code"),
                        "billed_amount": billed_amount,
                        "paid_amount": paid_amount,
                        "variance_amount": variance,
                        "match_status": status,
                        "reason_code": reason_code,
                    }
                )
            elif len(candidates) == 0:
                line_results.append(
                    {
                        "account_id": claim_id,
                        "claim_id": claim_id,
                        "dos_from": billed.get("dos_from"),
                        "dos_to": billed.get("dos_to"),
                        "proc_code": billed.get("proc_code"),
                        "billed_amount": _to_decimal(billed.get("billed_amount")),
                        "paid_amount": None,
                        "variance_amount": None,
                        "match_status": "unmatched_billed",
                        "reason_code": "missing_era_line",
                    }
                )
            else:
                collisions += 1
                line_results.append(
                    {
                        "account_id": claim_id,
                        "claim_id": claim_id,
                        "dos_from": billed.get("dos_from"),
                        "dos_to": billed.get("dos_to"),
                        "proc_code": billed.get("proc_code"),
                        "billed_amount": _to_decimal(billed.get("billed_amount")),
                        "paid_amount": None,
                        "variance_amount": None,
                        "match_status": "needs_review",
                        "reason_code": "collision",
                    }
                )

        for idx, era in enumerate(era_group):
            if idx in used_era:
                continue
            line_results.append(
                {
                    "account_id": claim_id,
                    "claim_id": claim_id,
                    "dos_from": era.get("dos_from"),
                    "dos_to": era.get("dos_to"),
                    "proc_code": era.get("proc_code"),
                    "billed_amount": None,
                    "paid_amount": _to_decimal(era.get("paid_amount")),
                    "variance_amount": None,
                    "match_status": "unmatched_era",
                    "reason_code": era.get("adj_code") or "unmatched_era",
                }
            )

        billed_total = Decimal("0")
        for row in billed_group:
            billed_total = _safe_add(billed_total, _to_decimal(row.get("billed_amount")))
        paid_total = Decimal("0")
        denial_seen = False
        for row in era_group:
            paid_total = _safe_add(paid_total, _to_decimal(row.get("paid_amount")))
            if _is_denial_code(row.get("adj_code")):
                denial_seen = True
        variance_total = paid_total - billed_total

        if billed_group and not era_group:
            claim_status = "PENDING_NO_ERA"
            reason_code = "missing_era_claim"
        elif era_group and not billed_group:
            claim_status = "NEEDS_REVIEW"
            reason_code = "UNMATCHED_ERA"
        else:
            if paid_total == Decimal("0") and denial_seen:
                claim_status = "DENIED"
                reason_code = "denial_adjustment_code"
            elif paid_total + tolerance < billed_total:
                claim_status = "UNDERPAID"
                reason_code = "paid_less_than_billed"
            elif abs(variance_total) <= tolerance:
                claim_status = "PAID"
                reason_code = None
            else:
                claim_status = "NEEDS_REVIEW"
                reason_code = "fallback_review"
            if collisions > 0 and claim_status == "PAID":
                claim_status = "NEEDS_REVIEW"
                reason_code = "collision"

        claim_results.append(
            {
                "account_id": claim_id,
                "claim_id": claim_id,
                "match_status": claim_status,
                "billed_total": billed_total if billed_group else None,
                "paid_total": paid_total if era_group else None,
                "variance_total": variance_total if billed_group and era_group else None,
                "line_count": max(len(billed_group), len(era_group)),
                "reason_code": reason_code,
            }
        )

    # Missing-key fallback matching.
    used_era_missing: set[int] = set()
    for billed in billed_missing:
        candidates = [
            idx
            for idx, era in enumerate(era_missing)
            if idx not in used_era_missing and _same_service(billed, era, tolerance)
        ]
        billed_amount = _to_decimal(billed.get("billed_amount"))
        if len(candidates) == 1:
            idx = candidates[0]
            used_era_missing.add(idx)
            era = era_missing[idx]
            paid_amount = _to_decimal(era.get("paid_amount"))
            variance = None
            if billed_amount is not None and paid_amount is not None:
                variance = paid_amount - billed_amount
            line_status = "matched" if variance is not None and abs(variance) <= tolerance else "needs_review"
            reason = era.get("adj_code") if line_status == "matched" else "missing_key_fallback"
            line_results.append(
                {
                    "account_id": None,
                    "claim_id": None,
                    "dos_from": billed.get("dos_from") or era.get("dos_from"),
                    "dos_to": billed.get("dos_to") or era.get("dos_to"),
                    "proc_code": billed.get("proc_code") or era.get("proc_code"),
                    "billed_amount": billed_amount,
                    "paid_amount": paid_amount,
                    "variance_amount": variance,
                    "match_status": line_status,
                    "reason_code": reason,
                }
            )
        elif len(candidates) == 0:
            line_results.append(
                {
                    "account_id": None,
                    "claim_id": None,
                    "dos_from": billed.get("dos_from"),
                    "dos_to": billed.get("dos_to"),
                    "proc_code": billed.get("proc_code"),
                    "billed_amount": billed_amount,
                    "paid_amount": None,
                    "variance_amount": None,
                    "match_status": "unmatched_billed",
                    "reason_code": "missing_key_no_era_match",
                }
            )
        else:
            line_results.append(
                {
                    "account_id": None,
                    "claim_id": None,
                    "dos_from": billed.get("dos_from"),
                    "dos_to": billed.get("dos_to"),
                    "proc_code": billed.get("proc_code"),
                    "billed_amount": billed_amount,
                    "paid_amount": None,
                    "variance_amount": None,
                    "match_status": "needs_review",
                    "reason_code": "missing_key_collision",
                }
            )

    for idx, era in enumerate(era_missing):
        if idx in used_era_missing:
            continue
        line_results.append(
            {
                "account_id": None,
                "claim_id": None,
                "dos_from": era.get("dos_from"),
                "dos_to": era.get("dos_to"),
                "proc_code": era.get("proc_code"),
                "billed_amount": None,
                "paid_amount": _to_decimal(era.get("paid_amount")),
                "variance_amount": None,
                "match_status": "unmatched_era",
                "reason_code": era.get("adj_code") or "unmatched_era_missing_key",
            }
        )

    # Build claim-level rows for missing-account line-level outputs.
    for line in [row for row in line_results if not (row.get("claim_id") or row.get("account_id") or "").strip()]:
        billed_total = _to_decimal(line.get("billed_amount"))
        paid_total = _to_decimal(line.get("paid_amount"))
        variance_total = _to_decimal(line.get("variance_amount"))
        status_hint = line.get("match_status")
        reason_code = line.get("reason_code")

        if status_hint == "unmatched_billed":
            claim_status = "PENDING_NO_ERA"
        elif status_hint == "unmatched_era":
            claim_status = "NEEDS_REVIEW"
            reason_code = reason_code or "UNMATCHED_ERA"
        else:
            if paid_total == Decimal("0") and _is_denial_code(reason_code):
                claim_status = "DENIED"
            elif billed_total is not None and paid_total is not None and paid_total + tolerance < billed_total:
                claim_status = "UNDERPAID"
            elif variance_total is not None and abs(variance_total) <= tolerance:
                claim_status = "PAID"
            else:
                claim_status = "NEEDS_REVIEW"

        claim_results.append(
            {
                "account_id": None,
                "claim_id": None,
                "match_status": claim_status,
                "billed_total": billed_total,
                "paid_total": paid_total,
                "variance_total": variance_total,
                "line_count": 1,
                "reason_code": reason_code,
            }
        )

    for claim in claim_results:
        status_value = claim.get("match_status")
        reason = _norm_text(claim.get("reason_code"))
        billed_total = _to_decimal(claim.get("billed_total"))
        paid_total = _to_decimal(claim.get("paid_total"))

        if status_value == "PAID":
            summary.matched_claims += 1
        elif status_value == "PENDING_NO_ERA":
            summary.unmatched_billed_claims += 1
        elif status_value == "UNDERPAID":
            summary.underpaid_claims += 1
        elif status_value == "DENIED":
            summary.denied_claims += 1
        elif status_value == "NEEDS_REVIEW":
            summary.needs_review_claims += 1
            if reason == "UNMATCHED_ERA" or (billed_total in {None, Decimal("0")} and (paid_total or Decimal("0")) > 0):
                summary.unmatched_era_claims += 1

    return claim_results, line_results, summary


def _write_output_xlsx(
    *,
    job: ReconImportJob,
    claim_results: list[dict[str, Any]],
    line_results: list[dict[str, Any]],
) -> str:
    output_rel = Path("outputs") / "recon" / job.org_id / f"{job.id}__recon.xlsx"
    output_path = _repo_root() / output_rel
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_claim = wb.active
    ws_claim.title = "ClaimResults"
    claim_headers = [
        "Claim ID",
        "Match Status",
        "Billed Total",
        "Paid Total",
        "Variance Total",
        "Line Count",
        "Reason Code",
    ]
    ws_claim.append(claim_headers)
    for row in claim_results:
        ws_claim.append(
            [
                row.get("claim_id") or row.get("account_id"),
                row.get("match_status"),
                float(row["billed_total"]) if row.get("billed_total") is not None else None,
                float(row["paid_total"]) if row.get("paid_total") is not None else None,
                float(row["variance_total"]) if row.get("variance_total") is not None else None,
                row.get("line_count"),
                row.get("reason_code"),
            ]
        )

    ws_line = wb.create_sheet("LineResults")
    line_headers = [
        "Claim ID",
        "DOS From",
        "DOS To",
        "Procedure",
        "Billed Amount",
        "Paid Amount",
        "Variance Amount",
        "Match Status",
        "Reason Code",
    ]
    ws_line.append(line_headers)
    for row in line_results:
        ws_line.append(
            [
                row.get("claim_id") or row.get("account_id"),
                row["dos_from"].isoformat() if isinstance(row.get("dos_from"), date) else None,
                row["dos_to"].isoformat() if isinstance(row.get("dos_to"), date) else None,
                row.get("proc_code"),
                float(row["billed_amount"]) if row.get("billed_amount") is not None else None,
                float(row["paid_amount"]) if row.get("paid_amount") is not None else None,
                float(row["variance_amount"]) if row.get("variance_amount") is not None else None,
                row.get("match_status"),
                row.get("reason_code"),
            ]
        )

    for ws in (ws_claim, ws_line):
        for col_idx, _ in enumerate(ws[1], start=1):
            max_len = 12
            col_cells = next(ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row))
            for cell in col_cells:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 42)

    wb.save(output_path)
    return output_rel.as_posix()


def _analyze_pdf(client: Any, model_id: str, pdf_path: Path) -> tuple[str, int, int, dict]:
    with pdf_path.open("rb") as f:
        poller = client.begin_analyze_document(model_id=model_id, body=f)
        result = poller.result()

    content = (getattr(result, "content", "") or "").strip()
    pages_detected = len(list(getattr(result, "pages", []) or []))
    tables_detected = len(list(getattr(result, "tables", []) or []))
    raw_json: dict = {}
    try:
        if hasattr(result, "to_dict"):
            raw_json = result.to_dict() or {}
        else:
            raw_json = json.loads(json.dumps(result, default=lambda o: o.__dict__))
    except Exception:
        raw_json = {"content": content}
    return content, pages_detected, tables_detected, raw_json


def _claim_next_job_id(db: Session) -> str | None:
    query = (
        select(ReconImportJob)
        .where(ReconImportJob.status == "queued")
        .order_by(ReconImportJob.created_at.asc())
        .limit(1)
    )
    if db.bind is not None and db.bind.dialect.name == "postgresql":
        query = query.with_for_update(skip_locked=True)

    try:
        row = db.execute(query).scalar_one_or_none()
    except OperationalError as exc:
        raise RuntimeError("recon tables missing; run database migrations before starting worker") from exc
    if row is None:
        return None

    row.status = "processing"
    row.started_at = utc_now()
    row.finished_at = None
    row.error_message = None
    db.add(row)
    db.commit()
    return row.id


def _build_claim_count(era_rows: list[dict[str, Any]], era_counters: dict[str, int]) -> int:
    keys: set[tuple[str | None, str | None]] = set()
    for row in era_rows:
        account_id = _claim_id(row)
        claim_no = (row.get("payer_claim_number") or "").strip() or None
        if account_id or claim_no:
            keys.add((account_id, claim_no))
    if keys:
        return len(keys)
    return int(era_counters.get("claim_blocks_found", 0))


def _process_job(job_id: str, *, tolerance: Decimal) -> None:
    db = SessionLocal()
    try:
        job = db.get(ReconImportJob, job_id)
        if job is None:
            return
        if job.status != "processing":
            return

        load_repo_dotenv()
        client, cfg = create_document_intelligence_client()
        model_id = (os.getenv("AZURE_DOCINTEL_MODEL") or cfg.model_id or "prebuilt-layout").strip() or "prebuilt-layout"

        actor_user = db.get(User, job.uploaded_by_user_id)
        actor_email = actor_user.email if actor_user else None

        log_event(
            db,
            action="recon_import_started",
            entity_type="recon_import_job",
            entity_id=job.id,
            organization_id=job.org_id,
            actor=actor_email,
            metadata={"job_id": job.id},
        )

        era_path = _repo_root() / job.era_storage_path
        billed_path = _repo_root() / job.billed_storage_path
        if not era_path.exists() or not billed_path.exists():
            raise FileNotFoundError("input_pdf_missing")

        era_content, pages_era, tables_era, era_raw = _analyze_pdf(client, model_id, era_path)
        billed_content, pages_billed, _tables_billed, billed_raw = _analyze_pdf(client, model_id, billed_path)

        db.add(
            DocumentAnalysis(
                id=str(uuid4()),
                org_id=job.org_id,
                job_id=job.id,
                document_type=DocumentType.ERA,
                raw_json=era_raw,
            )
        )
        db.add(
            DocumentAnalysis(
                id=str(uuid4()),
                org_id=job.org_id,
                job_id=job.id,
                document_type=DocumentType.BILLED,
                raw_json=billed_raw,
            )
        )

        era_rows: list[dict[str, Any]] = []
        billed_rows: list[dict[str, Any]] = []
        era_counters: dict[str, int] = {}
        billed_counters: dict[str, int] = {}

        era_rows, era_counters = parse_era_content(era_content, job_id=job.id)
        billed_rows, billed_counters = parse_billed_content(billed_content, billed_track="Billing")

        claim_results, line_results, recon_summary = _reconcile_rows(
            era_rows=era_rows,
            billed_rows=billed_rows,
            tolerance=tolerance,
        )

        db.execute(delete(EraLine).where(EraLine.job_id == job.id))
        db.execute(delete(BilledLine).where(BilledLine.job_id == job.id))
        db.execute(delete(ReconClaimResult).where(ReconClaimResult.job_id == job.id))
        db.execute(delete(ReconLineResult).where(ReconLineResult.job_id == job.id))

        era_orm_rows: list[EraLine] = []
        for row in era_rows:
            era_orm_rows.append(
                    EraLine(
                        job_id=job.id,
                        org_id=job.org_id,
                        account_id=row.get("account_id"),
                        payer_claim_number=row.get("payer_claim_number"),
                        icn=row.get("icn"),
                        member_id=row.get("member_id"),
                        dos_from=row.get("dos_from"),
                        dos_to=row.get("dos_to"),
                        proc_code=row.get("proc_code"),
                    units=_to_decimal(row.get("units")),
                    billed_amount=_to_decimal(row.get("billed_amount")),
                    allowed_amount=_to_decimal(row.get("allowed_amount")),
                    paid_amount=_to_decimal(row.get("paid_amount")),
                    adj_code=row.get("adj_code"),
                    adj_amount=_to_decimal(row.get("adj_amount")),
                    source_layout=row.get("source_layout"),
                )
            )
        if era_orm_rows:
            db.bulk_save_objects(era_orm_rows)

        billed_orm_rows: list[BilledLine] = []
        for row in billed_rows:
            billed_orm_rows.append(
                    BilledLine(
                        job_id=job.id,
                        org_id=job.org_id,
                        account_id=row.get("account_id"),
                        member_id=row.get("member_id"),
                        dos_from=row.get("dos_from"),
                        dos_to=row.get("dos_to"),
                        proc_code=row.get("proc_code"),
                    units=_to_decimal(row.get("units")),
                    billed_amount=_to_decimal(row.get("billed_amount")),
                )
            )
        if billed_orm_rows:
            db.bulk_save_objects(billed_orm_rows)

        claim_orm_rows: list[ReconClaimResult] = []
        for row in claim_results:
            claim_orm_rows.append(
                ReconClaimResult(
                    job_id=job.id,
                    org_id=job.org_id,
                    account_id=row.get("account_id"),
                    match_status=row.get("match_status") or "NEEDS_REVIEW",
                    billed_total=_to_decimal(row.get("billed_total")),
                    paid_total=_to_decimal(row.get("paid_total")),
                    variance_total=_to_decimal(row.get("variance_total")),
                    line_count=row.get("line_count"),
                    reason_code=row.get("reason_code"),
                )
            )
        if claim_orm_rows:
            db.bulk_save_objects(claim_orm_rows)

        line_orm_rows: list[ReconLineResult] = []
        for row in line_results:
            line_orm_rows.append(
                ReconLineResult(
                    job_id=job.id,
                    org_id=job.org_id,
                    account_id=row.get("account_id"),
                    dos_from=row.get("dos_from"),
                    dos_to=row.get("dos_to"),
                    proc_code=row.get("proc_code"),
                    billed_amount=_to_decimal(row.get("billed_amount")),
                    paid_amount=_to_decimal(row.get("paid_amount")),
                    variance_amount=_to_decimal(row.get("variance_amount")),
                    match_status=row.get("match_status") or "needs_review",
                    reason_code=row.get("reason_code"),
                )
            )
        if line_orm_rows:
            db.bulk_save_objects(line_orm_rows)

        output_xlsx_path = _write_output_xlsx(
            job=job,
            claim_results=claim_results,
            line_results=line_results,
        )

        job.status = "completed"
        job.pages_detected_era = pages_era
        job.tables_detected_era = tables_era
        job.claims_extracted_era = _build_claim_count(era_rows, era_counters)
        job.lines_extracted_era = len(era_rows)
        job.pages_detected_billed = pages_billed
        job.lines_extracted_billed = len(billed_rows)
        job.skipped_counts_json = {
            "era": {k: int(v) for k, v in era_counters.items()},
            "billed": {k: int(v) for k, v in billed_counters.items()},
        }
        job.matched_claims = recon_summary.matched_claims
        job.unmatched_era_claims = recon_summary.unmatched_era_claims
        job.unmatched_billed_claims = recon_summary.unmatched_billed_claims
        job.underpaid_claims = recon_summary.underpaid_claims
        job.denied_claims = recon_summary.denied_claims
        job.needs_review_claims = recon_summary.needs_review_claims
        job.output_xlsx_path = output_xlsx_path
        job.error_message = None
        job.finished_at = utc_now()

        db.add(job)
        db.commit()

        log_event(
            db,
            action="recon_import_completed",
            entity_type="recon_import_job",
            entity_id=job.id,
            organization_id=job.org_id,
            actor=actor_email,
            metadata={
                "job_id": job.id,
                "claims": len(claim_results),
                "lines": len(line_results),
                "matched_claims": recon_summary.matched_claims,
                "unmatched_billed_claims": recon_summary.unmatched_billed_claims,
                "unmatched_era_claims": recon_summary.unmatched_era_claims,
                "underpaid_claims": recon_summary.underpaid_claims,
                "denied_claims": recon_summary.denied_claims,
                "needs_review_claims": recon_summary.needs_review_claims,
            },
        )
    except Exception as exc:
        db.rollback()
        job = db.get(ReconImportJob, job_id)
        if job is not None:
            actor_user = db.get(User, job.uploaded_by_user_id)
            actor_email = actor_user.email if actor_user else None

            job.status = "failed"
            job.finished_at = utc_now()
            job.error_message = _sanitize_error_message(exc)
            db.add(job)
            db.commit()

            log_event(
                db,
                action="recon_import_failed",
                entity_type="recon_import_job",
                entity_id=job.id,
                organization_id=job.org_id,
                actor=actor_email,
                metadata={"job_id": job.id},
            )
        logger.exception("recon_import_worker_failed job_id=%s", job_id)
    finally:
        db.close()


def run_once(*, claim_limit: int = CLAIM_LIMIT) -> int:
    tolerance = RECON_TOLERANCE

    processed = 0
    for _ in range(max(1, claim_limit)):
        db = SessionLocal()
        try:
            job_id = _claim_next_job_id(db)
        finally:
            db.close()
        if not job_id:
            break
        _process_job(job_id, tolerance=tolerance)
        processed += 1
    return processed


def run_loop() -> None:
    logger.info("recon_import_worker_start poll_seconds=%s", POLL_SECONDS)
    while True:
        try:
            processed = run_once(claim_limit=CLAIM_LIMIT)
            if processed:
                logger.info("recon_import_worker_processed count=%s", processed)
        except Exception:
            logger.exception("recon_import_worker_loop_error")
        time.sleep(POLL_SECONDS)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Billing recon import worker")
    parser.add_argument("--once", action="store_true", help="Process currently queued jobs once then exit")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.once:
        processed = run_once(claim_limit=1000)
        logger.info("recon_import_worker_once_processed count=%s", processed)
        return 0

    run_loop()
    return 0


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    raise SystemExit(main())
